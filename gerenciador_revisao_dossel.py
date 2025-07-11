# 📌 Gerenciador de Revisão Dossel ajustado (controle de status e categorias)
import os
import subprocess
import sys
import smtplib
import time
from multiprocessing import Pool
from email.message import EmailMessage
from openpyxl import load_workbook
from tqdm import tqdm
from pathlib import Path
import json
from dotenv import load_dotenv
import streamlit as st

# Carrega variáveis de ambiente
load_dotenv()

# Configurações gerais
try:
    api_key = st.secrets["OPENAI_API_KEY"]
    id_bibliografico = st.secrets["ASSISTENTE_BIBLIOGRAFICO"]
    id_textual = st.secrets["ASSISTENTE_REVISOR_TEXTUAL"]
    id_tecnico = st.secrets["ASSISTENTE_TECNICO"]
    SENHA_APP = st.secrets["SENHA_APP"]
except ImportError:
    load_dotenv()
    api_key = os.getenv("OPENAI_API_KEY")
    id_bibliografico = os.getenv("ASSISTENTE_BIBLIOGRAFICO")
    id_textual = os.getenv("ASSISTENTE_REVISOR_TEXTUAL")
    id_tecnico = os.getenv("ASSISTENTE_TECNICO")
    SENHA_APP = os.getenv("SENHA_APP")

PASTA_ENTRADA = "entrada"
PASTA_SAIDA = "saida"
ARQUIVO_LOG_PROCESSADOS = "documentos_processados.txt"
ARQUIVO_LOG_FALHADOS = "documentos_falhados.txt"
STATUS_PATH_GLOBAL = "status.txt"

EMAIL_REMETENTE = os.getenv("EMAIL_REMETENTE") or "danilocnoleto952@gmail.com"
EMAIL_DESTINO = os.getenv("EMAIL_DESTINO") or "n.danilo@dosselambiental.com.br"

COTACAO_DOLAR = 5.65
MODO_LITE = "--lite" in sys.argv
TIMEOUT_LITE = 300

# Validação de documentos
def eh_documento_valido(nome_arquivo):
    return (
        nome_arquivo.endswith(".docx")
        and not nome_arquivo.startswith("~$")
        and not nome_arquivo.startswith("$")
        and not nome_arquivo.startswith("~")
    )

# Atualiza arquivo de status para barra de progresso
def atualizar_status_global(etapa):
    try:
        with open(STATUS_PATH_GLOBAL, "w", encoding="utf-8") as f:
            f.write(str(etapa))
    except Exception as e:
        print(f"Erro ao atualizar status global: {e}")

# Envia e-mail ao final
def enviar_email_final(resumo):
    msg = EmailMessage()
    msg["Subject"] = "✅ Revisão automática concluída"
    msg["From"] = EMAIL_REMETENTE
    msg["To"] = EMAIL_DESTINO
    msg.set_content("Todos os documentos foram processados com sucesso.\n\nResumo:\n\n" + resumo)
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(EMAIL_REMETENTE, SENHA_APP)
            smtp.send_message(msg)
        print("📬 E-mail de conclusão enviado com sucesso.")
    except Exception as e:
        print(f"❌ Erro ao enviar e-mail: {e}")

# Executa cada script secundário
def rodar_script(args):
    script, nome, usuario, entrada_path = args
    if not os.path.exists(script):
        print(f"❌ Script não encontrado: {script}")
        return 1
    comando = [sys.executable, script, str(entrada_path), usuario]
    if MODO_LITE and "revisor_dossel" in script:
        comando.append("--lite")
    print(f"▶️ Executando: {script} para {nome}")
    inicio = time.time()
    try:
        resultado = subprocess.run(comando, encoding="utf-8", errors="ignore")
    except subprocess.TimeoutExpired:
        print(f"⏱️ Timeout excedido: {script} para {nome}")
        registrar_falha(nome)
        return 1
    duracao = round(time.time() - inicio, 1)
    print(f"✅ Concluído {script} para {nome} em {duracao}s")
    pasta_saida_nome = os.path.join(PASTA_SAIDA, usuario, nome)
    tempo_path = os.path.join(pasta_saida_nome, f"tempo_{script.replace('.py','')}.txt")
    os.makedirs(os.path.dirname(tempo_path), exist_ok=True)
    with open(tempo_path, "w", encoding="utf-8") as f:
        f.write(f"{script} concluído em {duracao}s\n")
    if resultado.returncode != 0:
        log_path = os.path.join(pasta_saida_nome, f"erro_{script}.log")
        with open(log_path, "w", encoding="utf-8") as f:
            f.write(resultado.stdout + "\n" + resultado.stderr)
        registrar_falha(nome)
        return 1
    return 0

# Processa lotes via pool
def processar_assistente(script, nomes, usuario, etapa=None):
    if not nomes:
        return
    print(f"🚀 Iniciando {script} para {len(nomes)} documentos.")
    args = [(script, nome, usuario, entrada_path) for nome, entrada_path in nomes]
    with Pool(processes=min(3, len(nomes))) as pool:
        for nome, _ in zip(nomes, pool.imap_unordered(rodar_script, args)):
            if etapa:
                atualizar_status_global(etapa)

# Verifica se a revisão já gerou todos os arquivos
def documento_finalizado(nome, usuario):
    pasta = os.path.join(PASTA_SAIDA, usuario, nome)
    arquivos = [
        os.path.join(pasta, nome + "_revisado_completo.docx"),
        os.path.join(pasta, "relatorio_tecnico_" + nome + ".docx"),
        os.path.join(pasta, "avaliacao_completa.xlsx"),
    ]
    return all(os.path.exists(a) for a in arquivos)

# Registra documentos processados ou com falha
def registrar_documento(nome):
    with open(ARQUIVO_LOG_PROCESSADOS, "a", encoding="utf-8") as f:
        f.write(nome + "\n")

def registrar_falha(nome):
    with open(ARQUIVO_LOG_FALHADOS, "a", encoding="utf-8") as f:
        f.write(nome + "\n")

def ja_foi_processado(nome):
    if not os.path.exists(ARQUIVO_LOG_PROCESSADOS):
        return False
    with open(ARQUIVO_LOG_PROCESSADOS, "r", encoding="utf-8") as f:
        return nome in [l.strip() for l in f]

# Gera resumo de custos a partir dos XLSX
def gerar_resumo_custos():
    resumo = []
    for user_dir in os.listdir(PASTA_SAIDA):
        pasta_user = os.path.join(PASTA_SAIDA, user_dir)
        if not os.path.isdir(pasta_user):
            continue
        for pasta in os.listdir(pasta_user):
            path = os.path.join(pasta_user, pasta, "avaliacao_completa.xlsx")
            if not os.path.isfile(path):
                continue
            try:
                wb = load_workbook(path)
                ws = wb["Resumo"]
                ultima = ws.max_row
                tempo = ws.cell(row=ultima, column=2).value
                usd = ws.cell(row=ultima, column=4).value
                brl = ws.cell(row=ultima, column=5).value
                if usd and float(usd) > 0:
                    resumo.append(f"{pasta} → ⏱ {tempo}s | 💵 US$ {round(float(usd),4)} | R$ {round(float(brl),2)}")
            except:
                pass
    return "\n".join(resumo)

# Carrega documentos que falharam
def carregar_falhados():
    if not os.path.exists(ARQUIVO_LOG_FALHADOS):
        return []
    with open(ARQUIVO_LOG_FALHADOS, "r", encoding="utf-8") as f:
        return [l.strip() for l in f]

# Extrai categorias do JSON de mapeamento
def categorias_do_documento(nome, usuario):
    try:
        path_json = os.path.join(PASTA_SAIDA, usuario, nome, "mapeamento_textual.json")
        with open(path_json, encoding="utf-8") as f:
            dados = json.load(f)
        return {item.get("categoria") for item in dados if item.get("categoria")}
    except:
        return set()

# === Função principal ===
def main():
    retry = "--retry" in sys.argv
    if os.path.exists(STATUS_PATH_GLOBAL):
        os.remove(STATUS_PATH_GLOBAL)

    raw_args = [arg for arg in sys.argv[1:] if not arg.startswith("--")]
    if len(raw_args) < 2:
        print("❌ Parâmetros insuficientes.")
        return

    entrada_path = Path(raw_args[0]).resolve()
    usuario = raw_args[1]
    
    os.environ["USUARIO"] = usuario

    nomes = []
    if not retry:
        if entrada_path.is_file() and entrada_path.suffix == ".docx":
            nome = entrada_path.stem
            nomes.append((nome, entrada_path))
    else:
        nomes = carregar_falhados()

    if not nomes:
        print("📬 Nenhum documento novo para processar.")
        return

    print(f"🔎 Documentos a revisar: {len(nomes)}")
    processar_assistente("mapeador.py", nomes, usuario, etapa=25)

    for nome, entrada_path in nomes:
        categorias = categorias_do_documento(nome, usuario)
        print(f"📂 {nome}: Categorias detectadas: {sorted(categorias)}")

        # Força execução textual se nenhuma categoria for detectada
        if "textual" in categorias or not categorias:
            rodar_script(("revisor_dossel_v2_final.py", nome, usuario, entrada_path))
            atualizar_status_global(50)
        if "bibliografico" in categorias and not MODO_LITE:
            rodar_script(("verificador_bibliografico_final.py", nome, usuario, entrada_path))
            atualizar_status_global(75)
        if not MODO_LITE:
            rodar_script(("revisor_falhas.py", nome, usuario, entrada_path))
            atualizar_status_global(100)
        else:
            atualizar_status_global(100)

    for nome, _ in nomes:
        if documento_finalizado(nome, usuario):
            registrar_documento(nome)

    if not MODO_LITE:
        resumo = gerar_resumo_custos()
        print("\n📊 Resumo de custos:\n" + resumo)
        enviar_email_final(resumo)

if __name__ == "__main__":
    inicio = time.time()
    main()
    atualizar_status_global(100)
    print(f"\n⏱️ Tempo total: {round(time.time() - inicio, 1)}s")
