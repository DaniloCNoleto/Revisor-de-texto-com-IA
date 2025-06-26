# ✅ Etapa 1: Gerar resumo_final.json consolidado por documento 
import os
import json
import re
import time
from openpyxl import load_workbook

PASTA_SAIDA = "saida"

def consolidar_resumo(documento):
    pasta = os.path.join(PASTA_SAIDA, documento)
    resumo = {
        "documento": documento,
        "etapas": {},
        "falhas": {},
        "tempo (em s)": 0,
        "custo (US$)": 0.0,
        "custo (R$)": 0.0,
        "similaridades": {},
        "media_similaridade": 0.0
    }

    etapas = [
        ("textual", os.path.join(pasta, "checkpoint.txt")),
        ("bibliografica", os.path.join(pasta, "checkpoint_biblio.txt")),
        ("falhas", os.path.join(pasta, "checkpoint_falhas.txt")),
        ("trackchanges", os.path.join(pasta, documento + "_revisado e rastreado.docx"))
    ]
    for nome, path in etapas:
        resumo["etapas"][nome] = os.path.exists(path)

    for arquivo in os.listdir(pasta):
        if arquivo.startswith("tempo_") and arquivo.endswith(".txt"):
            try:
                with open(os.path.join(pasta, arquivo), encoding="utf-8") as f:
                    m = re.search(r"(\d+(\.\d+)?)s", f.read())
                    if m:
                        resumo["tempo (em s)"] += float(m.group(1))
            except:
                pass

    planilhas = ["avaliacao_quantitativa.xlsx", "avaliacao_bibliografica.xlsx", "avaliacao_falhas.xlsx"]
    for nome_arq in planilhas:
        caminho = os.path.join(pasta, nome_arq)
        if os.path.exists(caminho):
            try:
                wb = load_workbook(caminho)
                for ws in wb.worksheets:
                    if ws.title.lower() == "similaridade":
                        sim_values = [c.value for c in ws["B"] if isinstance(c.value, (int, float))]
                        if sim_values:
                            resumo["similaridades"][nome_arq.replace("avaliacao_", "").replace(".xlsx", "")] = round(sum(sim_values) / len(sim_values), 2)
                    for row in ws.iter_rows(min_row=ws.max_row-3, max_row=ws.max_row):
                        for cel in row:
                            if isinstance(cel.value, (int, float)):
                                if "usd" in str(row[0].value).lower():
                                    resumo["custo (US$)"] += float(cel.value)
                                if "brl" in str(row[0].value).lower():
                                    resumo["custo (R$)"] += float(cel.value)
            except:
                pass

    for etapa in ["falhas_textuais.json", "falhas_bibliograficas.json"]:
        caminho = os.path.join(pasta, etapa)
        if os.path.exists(caminho):
            try:
                falhas = json.load(open(caminho, encoding="utf-8"))
                resumo["falhas"][etapa] = len(falhas)
            except:
                resumo["falhas"][etapa] = "erro leitura"

    sim_vals = list(resumo["similaridades"].values())
    if sim_vals:
        resumo["media_similaridade"] = round(sum(sim_vals) / len(sim_vals), 2)

    json.dump(resumo, open(os.path.join(pasta, "resumo_final.json"), "w", encoding="utf-8"), indent=2, ensure_ascii=False)
    return resumo

# Exemplo de uso isolado:
if __name__ == "__main__":
    for doc in os.listdir(PASTA_SAIDA):
        if os.path.isdir(os.path.join(PASTA_SAIDA, doc)):
            consolidar_resumo(doc)
            print(f"🔍 Consolidado: {doc}")
