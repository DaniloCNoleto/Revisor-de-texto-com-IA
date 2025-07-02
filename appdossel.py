import streamlit as st
import os
import re
import shutil
import time
import subprocess
import sys
import sqlite3
import datetime
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
import plotly.express as px
from passlib.hash import pbkdf2_sha256
try:
    from passlib.hash import pbkdf2_sha256
except Exception:  # pragma: no cover - fallback when passlib missing
    pbkdf2_sha256 = None
from datetime import datetime
from urllib.parse import urlparse, parse_qs  # ➜ novo import para utilidades de URL
from streamlit_option_menu import option_menu
import sqlite3
from filelock import FileLock
import hashlib
import hmac
# --- Drive service-account ---
import json
import base64
import io
import mimetypes
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload


SA_INFO = json.loads(
    base64.b64decode(os.environ["SA_KEY_B64"]).decode()
)
CREDS = service_account.Credentials.from_service_account_info(
            SA_INFO,
            scopes=["https://www.googleapis.com/auth/drive.file"]
        )
DRIVE = build("drive", "v3", credentials=CREDS)
FOLDER_ID = os.environ["FOLDER_ID"]         # 1cN0r1gyy9kN2S7_n-5NmpVOzlnjRsStU


# ------------------------------------------------------------------
# ------------------------ Paths e DB ------------------------------
# ------------------------------------------------------------------

DB_PATH = Path("users.db")
DB_LOCK = FileLock(str(DB_PATH) + ".lock")
PASTA_ENTRADA = Path("entrada")
PASTA_SAIDA = "saida"
PASTA_HISTORICO = Path("historico")
STATUS_PATH = Path("status.txt")
LOG_PROCESSADOS = Path("documentos_processados.txt")
LOG_FALHADOS = Path("documentos_falhados.txt")
QUEUE_FILE = Path("queue.txt")

# --- Inicialização do DB ---

def init_db():
    with DB_LOCK:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                email TEXT UNIQUE NOT NULL,
                full_name TEXT,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        c.execute(
            """
            CREATE TABLE IF NOT EXISTS revisions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                file_name TEXT NOT NULL,
                processed_path TEXT NOT NULL,
                timestamp TEXT NOT NULL,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
            """
        )
        conn.commit()
        conn.close()

def upload_e_link(path: Path) -> str:
    """Envia <path> à pasta do Drive e devolve URL pública de download."""
    if not path.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")
    mime = mimetypes.guess_type(path)[0] or "application/octet-stream"

    # 1) upload
    meta = {"name": path.name, "parents": [FOLDER_ID]}
    try:
        media = MediaIoBaseUpload(path.open("rb"), mimetype=mime)
    except FileNotFoundError:
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")
    file = DRIVE.files().create(body=meta, media_body=media,
                                fields="id").execute()
    file_id = file["id"]

    # 2) libera “anyone with link” → read
    DRIVE.permissions().create(fileId=file_id,
                               body={"type": "anyone", "role": "reader"}
                               ).execute()

    # 3) link direto
    return f"https://drive.google.com/uc?export=download&id={file_id}"


# --- opcional: restaurar + backup do users.db ----------------------------
def restore_db():
    try:
        from googleapiclient.http import MediaIoBaseDownload
    except Exception as e:
        print("[restore_db] falha no import ➜", e)
        return
    res = DRIVE.files().list(
        q=f"'{FOLDER_ID}' in parents and name='users.db'",
        orderBy="modifiedTime desc",
        pageSize=1,
        fields="files(id)"
    ).execute()

    if not res.get("files"):
        print("[restore_db] nada encontrado")
        return

    file_id = res["files"][0]["id"]
    request = DRIVE.files().get_media(fileId=file_id)
    tmp_path = DB_PATH.with_suffix('.tmp')
    with DB_LOCK, io.FileIO(tmp_path, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
    os.replace(tmp_path, DB_PATH)

    os.chmod(DB_PATH, 0o666)
    print("[restore_db] baixado", DB_PATH.stat().st_size, "bytes")

def mark_db_dirty():
    # seta flag na sessão global
    st.session_state["db_dirty"] = True


def backup_db():
    # 0) roda só se mexemos no banco
    if not st.session_state.get("db_dirty"):
        print("[backup_db] ignorado – nada mudou")
        return

    # 1) garante commits encerrados
    with DB_LOCK:
        with sqlite3.connect(DB_PATH):
            pass

    # 2) procura users.db na pasta
        sha1_local = hashlib.sha1(DB_PATH.read_bytes()).hexdigest()

        res = DRIVE.files().list(
            q=f"'{FOLDER_ID}' in parents and name='users.db'",
            fields="files(id, appProperties)",
            pageSize=1
        ).execute()

        media = MediaIoBaseUpload(open(DB_PATH, "rb"),
                                  mimetype="application/octet-stream",
                                  resumable=True)

        tries = 3
        while tries:
            try:
                if res.get("files"):
                    fid = res["files"][0]["id"]
                    remote_sha = res["files"][0].get("appProperties", {}).get("sha1")
                    if remote_sha == sha1_local:
                        print("[backup_db] sem mudanças")
                        break
                    body = {"appProperties": {"sha1": sha1_local}}
                    DRIVE.files().update(fileId=fid, media_body=media, body=body).execute()
                    print("[backup_db] users.db atualizado")
                else:
                    body = {"name": "users.db", "parents": [FOLDER_ID], "appProperties": {"sha1": sha1_local}}
                    DRIVE.files().create(body=body, media_body=media).execute()
                    print("[backup_db] users.db criado")
                break
            except Exception as e:
                tries -= 1
                if not tries:
                    print(f"[backup_db] falhou: {e}")
                    return
                sleep_time = 2 ** (3 - tries)
                print(f"[backup_db] erro, retry em {sleep_time}s")
                time.sleep(sleep_time)

    st.session_state["db_dirty"] = False

    # 2) garante permissão pública (caso tenha sido criado agora)
    #    — opcional, mas útil se quiser baixar manualmente sem login —
    # DRIVE.permissions().create(
    #     fileId=fid, body={"type": "anyone", "role": "reader"}
    # ).execute()



# --- Autenticação ---

def hash_password(password: str) -> str:
    return pbkdf2_sha256.hash(password)
    """Hash de senha com passlib ou fallback nativo."""

    if pbkdf2_sha256:
        return pbkdf2_sha256.hash(password)

    # Fallback simplificado em caso de ausência do passlib
    salt = os.urandom(12)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 29000)
    return "$pbkdf2-sha256$29000$" + base64.b64encode(salt).decode().rstrip("=") + "$" + \
        base64.b64encode(dk).decode().rstrip("=")

def verify_password(password: str, hash_str: str) -> bool:
    """Verifica senha utilizando passlib ou implementação local."""
    if pbkdf2_sha256:
        try:
            return pbkdf2_sha256.verify(password, hash_str)
        except Exception:
            return False

    try:
        if not hash_str.startswith("$pbkdf2-sha256$"):
            return False
        _, _, rounds, salt_b64, hash_b64 = hash_str.split("$")
        salt = base64.b64decode(salt_b64 + "=" * (-len(salt_b64) % 4))
        dk = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, int(rounds))
        calc = base64.b64encode(dk).decode().rstrip("=")
        return hmac.compare_digest(calc, hash_b64)
    except Exception:
        return False

def register_user(username: str, email: str, full_name: str, password: str) -> bool:
    with DB_LOCK:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        try:
            pwd_hash = hash_password(password)
            now = datetime.now().isoformat()

            c.execute(
                "INSERT INTO users (username, email, full_name, password_hash, created_at) VALUES (?, ?, ?, ?, ?)",
                (username, email, full_name, pwd_hash, now)
            )
            conn.commit()
            mark_db_dirty()
            return True
        except sqlite3.IntegrityError:
            return False
        finally:
            conn.close()


def authenticate_user(username: str, password: str) -> dict | None:
    with DB_LOCK:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT id, password_hash, full_name FROM users WHERE username = ?", (username,))
        row = c.fetchone()
        conn.close()
    if row and verify_password(password, row[1]):
        return {"id": row[0], "username": username, "full_name": row[2]}
    return None

# --- Histórico de Revisões ---

def log_revision(
    user_id: int,
    file_name: str,
    processed_path: str,
    timestamp: str | None = None,
):
    """Registra uma revisão no banco de dados.

    Quando `timestamp é fornecido, ele permite agrupar registros    relacionados (por exemplo, documento e relatório) pelo mesmo instante.
    """
    with DB_LOCK:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        now = timestamp or datetime.now().isoformat()
        c.execute(
            "INSERT INTO revisions (user_id, file_name, processed_path, timestamp) VALUES (?, ?, ?, ?)",
            (user_id, file_name, processed_path, now)
        )
        conn.commit()
        conn.close()
        mark_db_dirty()


def get_history(user_id: int) -> list[tuple[str, str, str]]:
    try:
        with DB_LOCK:
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute(
                """
                SELECT file_name, processed_path, timestamp
                FROM revisions
                WHERE user_id = ?
                ORDER BY timestamp DESC
                """,
                (user_id,)
            )
            rows = c.fetchall()
            return rows
    except sqlite3.Error as e:
        print(f"❌ Erro no banco de dados: {e}")
        return []
    finally:
        if 'conn' in locals():
            conn.close()


# --- Fila e status ---

def load_queue():
    if QUEUE_FILE.exists():
        return [l.strip() for l in QUEUE_FILE.read_text().splitlines() if l.strip()]
    return []

def save_queue(q):
    QUEUE_FILE.write_text("\n".join(q))

def add_to_queue(nome):
    q = load_queue()
    if nome not in q:
        q.append(nome)
        save_queue(q)
    return q.index(nome) + 1

def remove_from_queue(nome):
    q = load_queue()
    if nome in q:
        q.remove(nome)
        save_queue(q)


# --- CSS e Layout ---
def apply_css() -> None:
    st.markdown(
        """
        <style>
        /* ---------- Paleta Dossel ---------- */
        :root {
            --dossel-green-600: #007f56;
            --dossel-green-700: #005f43;
            --dossel-green-400: #00AF74;
            --dossel-green-100: #95afc0;
            --background-color: #fff;
            --text-color: #000;
            --sidebar-bg-light: var(--dossel-green-100);

        }
        @media (prefers-color-scheme: dark) {
            :root {
                --background-color: #111;
                --text-color: #eee;
                --sidebar-bg: rgba(0, 127, 86, .15);
            }
        }

        html[data-theme="dark"] {
            --background-color: #111;
            --text-color: #eee;
            --sidebar-bg: rgba(0, 127, 86, .15);
            }
        /* ---------- Corrige o FUNDO BRANCO que sobrou ---------- */
        /* 1) contêiner principal da página */
        [data-testid="stAppViewContainer"],
        /* 2) bloco central (quando existir) */
        [data-testid="block-container"],
        /* 3) qualquer div de 1º nível dentro do contêiner (pega as css-hash) */
        [data-testid="stAppViewContainer"] > div {
            background: var(--background-color) !important;
            color: var(--text-color) !important;   /* força textos dentro */
        }

        /* zerar fundo de colunas e métricas */
        [data-testid="column"],
        [data-testid="metric-container"] {
            background: transparent !important;
            color: var(--text-color) !important;
        }

        /* ---------- Elementos globais ---------- */
        html, body, .stApp, [class*="css"] {
            background: var(--background-color) !important;
            color: var(--text-color) !important;
            font-family: 'Inter', sans-serif;
        }

        .stApp > header, .stApp > footer { display: none !important; }

        /* ---------- Identidade Dossel (igual à versão anterior) ---------- */
        .title-dossel { color: var(--dossel-green-600); }
        html[data-theme="dark"] .title-dossel,
        body[data-theme="dark"]  .title-dossel { color: var(--dossel-green-400); }

        .stButton button {
            background: var(--dossel-green-600) !important;
            color: #fff !important;
        }
        .stButton button:hover { background: var(--dossel-green-700) !important; }

        .stDownloadButton button {
            background: transparent !important;
            border: 2px solid var(--dossel-green-600) !important;
            color: var(--dossel-green-600) !important;
        }
        .stDownloadButton button:hover {
            background: var(--dossel-green-400) !important;
            border-color: var(--dossel-green-400) !important;
            color: #fff !important;
        }
        .stButton, .stDownloadButton, .stLinkButton { display:flex; justify-content:center; }
        .stButton button, .stDownloadButton button, .stLinkButton a { margin:auto; max-width:320px; }

        .main-box { display:flex; flex-direction:column; align-items:center; text-align:center; gap:1rem; }

        /* ---------- Sidebar ---------- */
        section[data-testid="stSidebar"] > div:first-child {
            background: var(--sidebar-bg-light) !important;
            padding-top: 2rem;
        }
        html[data-theme="dark"] section[data-testid="stSidebar"] > div:first-child,
        body[data-theme="dark"] section[data-testid="stSidebar"] > div:first-child {
            background: var(--sidebar-bg) !important;
        }

        /* ---------- CENTRALIZA TODO O CONTEÚDO ---------- */
        /* o Streamlit envolve cada página em data-testid="block-container" */
        [data-testid="block-container"] {
            margin: 14px auto 0;
            padding-left: .5rem;
            padding-right: .5rem;
            display:flex;
            flex-direction:column;
            align-items:center;
            min-height:calc(100vh - 4rem);
            justify-content:center;
        }

        </style>
        """,
        unsafe_allow_html=True,
    )

    # ---------- lógica para redirecionar quem ainda não fez login ----------
    if "user" not in st.session_state:
        st.session_state["pagina"] = "login"
        header()
        page_login()
        footer()
        st.stop()


def header():
    st.markdown('<div class="main-box">', unsafe_allow_html=True)
    logo_path = Path("Dossel - Logo Horizontal.png")
    if logo_path.exists():
        img_b64 = base64.b64encode(logo_path.read_bytes()).decode()
        st.markdown(
            f'<div class="logo-dossel"><img src="data:image/png;base64,{img_b64}" alt="Logo Dossel"></div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            '<div class="logo-dossel">'
            '  <img src="Dossel - Logo Horizontal.png" '
            '       alt="Logo Dossel">'
            '</div>',
            unsafe_allow_html=True,
        )
    st.markdown('<div class="title-dossel">Revisor Automático Dossel</div>', unsafe_allow_html=True)


def page_login():
    st.markdown('<div class="main-box">', unsafe_allow_html=True)
    st.subheader("Login")
    username = st.text_input("Usuário", key="login_username")
    password = st.text_input("Senha", type="password", key="login_password")
    if st.button("Entrar", key="login_enter"):
        user = authenticate_user(username, password)
        if user:
            st.session_state.clear()
            st.session_state['user'] = user
            st.session_state['usuario'] = user['username']
            st.session_state['pagina'] = 'upload'
            st.rerun()
        else:
            st.error("Credenciais inválidas")
    st.markdown("---")
    st.write("Ainda não tem conta? ")
    if st.button("Registrar-se", key="login_register"):
        st.session_state['show_register'] = True

    # Se estiver pedindo registro
    if st.session_state.get('show_register'):
        page_register()

    st.markdown('</div>', unsafe_allow_html=True)


def page_register():
    st.markdown('<div class="main-box">', unsafe_allow_html=True)
    st.subheader("Registro de Usuário")
    new_user = st.text_input("Usuário", key="register_username")
    email = st.text_input("E-mail", key="register_email")
    full_name = st.text_input("Nome Completo", key="register_fullname")
    pwd = st.text_input("Senha", type="password", key="register_password")
    pwd2 = st.text_input("Confirme a Senha", type="password", key="register_password2")
    if st.button("Criar Conta", key="register_create"):
        if pwd != pwd2:
            st.error("Senhas não coincidem")
        elif register_user(new_user, email, full_name, pwd):
            st.success("Conta criada com sucesso! Faça login.")
            st.session_state.pop('show_register', None)
        else:
            st.error("Usuário ou e-mail já cadastrado")
    st.markdown('</div>', unsafe_allow_html=True)

# Page_history no appdossel.py com correção de chave duplicada e identificação de tipo de revisão

# helper: decide entre abrir link ou baixar arquivo local
# ───── Helper único para todos os downloads ────────────────────────────
def botao_download(label: str, destino: str | None, *, key: str):
    if not destino:
        return
    if destino.startswith(("http://", "https://")):
        st.link_button(label, url=destino, use_container_width=True)
    else:
        p = Path(destino)
        if p.exists():
            st.download_button(label, p.read_bytes(),
                               file_name=p.name, key=key)
        else:
            st.warning("⚠️ Arquivo não encontrado.")


# ───── Histórico agrupando relatórios e docs revisados ─────────────────
def page_history():
    st.subheader("Histórico de Revisões")

    user = st.session_state.get("user")
    if not user:
        st.error("Usuário não autenticado.")
        return

    usuario = user["username"]
    linhas = get_history(user["id"])      # (file_name, processed_path, ts_iso)
    if not linhas:
        st.info("Nenhuma revisão encontrada.")
        return

    # 1️⃣ agrupa por timestamp (documento + relatório com o mesmo instante)
    grupos: dict[tuple[str, str], dict] = {}
    independentes: list[tuple] = []

    for fname, pth, ts in linhas:
        if pth.startswith(("http://", "https://")):
            raiz = fname.removeprefix("Relatório ").strip()
            ts_key = ts.split(".")[0]  # ignora microssegundos
            chave = (raiz, ts_key)
            g = grupos.setdefault(chave, {"raiz": raiz, "doc": None, "rel": None, "data": ts})
            if fname.lower().startswith("relatório"):
                g["rel"] = pth
            else:
                g["doc"] = pth          # data principal = doc revisado
        else:
            independentes.append((fname, pth, ts))

    # 2️⃣ renderiza grupos (mais recentes primeiro)
    ordenados = sorted(grupos.values(), key=lambda x: x["data"], reverse=True)
    for info in ordenados:
        data_br = datetime.fromisoformat(info["data"]).strftime("%d/%m/%Y")
        st.write(f"**{data_br} — {info['raiz']}**")
        col1, col2 = st.columns(2)
        
        if info["doc"]:
            with col1:
                botao_download(
                    "📄 Download Revisado",
                    info["doc"],
                    key=f"{info['raiz']}_{info['data']}_doc",
                )
        if info["rel"]:
            with col2:
                botao_download(
                    "📑 Download Relatório",
                    info["rel"],
                    key=f"{info['raiz']}_{info['data']}_rel",
                )
        st.markdown("---")

    # 3️⃣ renderiza itens antigos (pasta local) — fluxo original ------------
    for fname, processed_path, ts_iso in independentes:
        data_br = datetime.fromisoformat(ts_iso).strftime("%d/%m/%Y")
        st.write(f"**{data_br} — {fname}**")

        dir_saida = Path(processed_path) if processed_path else None
        if not (dir_saida and dir_saida.exists()):
            dir_saida = Path(PASTA_SAIDA) / usuario / fname
        if not dir_saida.exists():
            candidatos = list(Path(PASTA_SAIDA).glob(f"*/{fname}"))
            if candidatos:
                dir_saida = candidatos[0]

        if not dir_saida.exists() or not dir_saida.is_dir():
            st.warning("⚠️ Pasta de saída não encontrada.")
            st.markdown("---")
            continue

        doc_final, relatorio, tipo = None, None, "Desconhecido"
        for child in dir_saida.iterdir():
            if child.suffix == ".docx":
                if "_revisado" in child.name and not doc_final:
                    doc_final = child
                    if "completo" in child.name:
                        tipo = "Revisão Completa"
                    elif "texto" in child.name:
                        tipo = "Revisão Rápida"
                    elif "falhas" in child.name:
                        tipo = "Revisão com Falhas"
                    elif "biblio" in child.name:
                        tipo = "Revisão Bibliográfica"
                    else:
                        tipo = "Revisado"
                elif child.name.startswith("relatorio_tecnico"):
                    relatorio = child

        st.caption(f"🧾 Tipo: {tipo}")
        col1, col2 = st.columns(2)
        if doc_final:
            with col1:
                botao_download("📄 Download Revisado", str(doc_final),
                               key=f"{fname}_{ts_iso}_doc")
        if relatorio:
            with col2:
                botao_download("📑 Download Relatório", str(relatorio),
                               key=f"{fname}_{ts_iso}_rel")
        st.markdown("---")


# --- Fluxo Original de Revisão ---
def page_upload():
    if st.session_state.get("pagina") != "upload":
        return

    # Limpa estados antigos que atrapalham a transição
    for key in ['modo_selected', 'modo_lite', 'removed_from_queue', 'want_start', 'processo_iniciado', 'entrada_path']:
        st.session_state.pop(key, None)

    st.subheader("Envie um arquivo .docx para revisão:")
    arquivo = st.file_uploader("Selecione um arquivo .docx para revisão:", type="docx", label_visibility='collapsed')

    if not arquivo:
        return

    nome = arquivo.name.replace('.docx', '')
    usuario = st.session_state.get('usuario')
    st.session_state['nome'] = nome
    st.session_state['usuario'] = usuario
    st.write(f"**Arquivo carregado:** {nome}")

    if st.button(f"▶️ Iniciar Revisão: {nome}"):
        st.session_state['want_start'] = True

    if st.session_state.get('want_start'):
        # Cria pasta de entrada específica do usuário
        pasta_entrada_usuario = PASTA_ENTRADA / usuario
        pasta_entrada_usuario.mkdir(parents=True, exist_ok=True)

        # Limpa arquivos anteriores
        for fpath in pasta_entrada_usuario.iterdir():
            try:
                if fpath.is_file():
                    fpath.unlink()
                elif fpath.is_dir():
                    shutil.rmtree(fpath)
            except FileNotFoundError:
                pass

        file_path = pasta_entrada_usuario / arquivo.name
        with open(file_path, 'wb') as f:
            f.write(arquivo.getbuffer())

        # Atualiza estado e avança para próxima página
        st.session_state['entrada_path'] = str(file_path)
        st.session_state['pagina'] = 'modo'
        st.rerun()


def page_mode():
    nome = st.session_state['nome']

    if not st.session_state.get('modo_selected'):
        st.markdown('### Escolha o tipo de revisão:')
        c1, c2 = st.columns(2)
        if c1.button('🔎 Revisão Completa'):
            st.session_state['modo_selected'] = True
            st.session_state['modo_lite'] = False
            st.rerun()
        if c2.button('⚡ Revisão Simples'):
            st.session_state['modo_selected'] = True
            st.session_state['modo_lite'] = True
            st.rerun()

    elif not st.session_state.get('pagina') == 'acompanhamento':  # <-- evita reexibir se já mudou
        modo = 'Lite' if st.session_state['modo_lite'] else 'Completa'
        st.markdown(f"#### Você quer realizar a revisão **{modo}** do documento **{nome}**?")
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button('✅ Confirmar Revisão'):
                st.session_state['pagina'] = 'acompanhamento'
                st.session_state['processo_iniciado'] = False
                st.rerun()
        with col2:
            if st.button('🔙 Voltar'):
                st.session_state['pagina'] = 'upload'
                st.rerun()

def page_progress():
    entrada_path = st.session_state.get("entrada_path")
    nome = st.session_state.get("nome")
    usuario = st.session_state.get("usuario")

    # ─── validações ───────────────────────────────────────────────
    if not entrada_path or not nome:
        st.session_state["pagina"] = "upload"
        st.rerun()

    lite = st.session_state.get("modo_lite", False)
    gerenciador = Path(__file__).parent / "gerenciador_revisao_dossel.py"

    # ─── dispara o gerenciador 1× só ─────────────────────────────
    if not st.session_state.get("processo_iniciado", False):

        # 0. limpa status
        STATUS_PATH.unlink(missing_ok=True)

        # 1. versiona revisão anterior (LOCAL → opcional upload)
        antiga = Path(PASTA_SAIDA) / usuario / nome
        if antiga.exists():
            user = st.session_state["user"]
            hist_dir = Path(PASTA_HISTORICO) / user["username"]
            hist_dir.mkdir(parents=True, exist_ok=True)

            versoes = [int(m.group(1))
                       for p in hist_dir.iterdir()
                       if (m := re.match(fr"^{re.escape(nome)}_v(\d+)$", p.name))]
            dest = hist_dir / f"{nome}_v{max(versoes, default=0)+1}"
            shutil.move(str(antiga), str(dest))


            # (a) sobe ZIP da pasta antiga (se quiser registrar no Drive)
            # link_ant = upload_e_link(shutil.make_archive(dest, "zip", dest))
            # (b) ou apenas guarda o caminho local
            log_revision(user["id"], nome, str(dest))
            backup_db()

        # 2. garante que o script existe
        if not gerenciador.exists():
            st.error(f"Script não encontrado: {gerenciador}")
            return

        # 3. chama o gerenciador (sub-processo)
        with st.spinner("👷 Iniciando gerenciador…"):
            subprocess.Popen(
                [sys.executable, str(gerenciador), entrada_path, usuario]
                + (["--lite"] if lite else [])
            )
        st.session_state["processo_iniciado"] = True
        st.rerun()

    # ─── barra de progresso ──────────────────────────────────────
    v = int(STATUS_PATH.read_text().strip()) if STATUS_PATH.exists() else 0
    st.markdown(
        f"""
        <div style="position: relative; width: 100%; background:#f0f0f0;
                    border-radius:4px;height:30px;margin:10px auto;">
          <div style="width:{v}%;background:#007f56;height:100%;
                      border-radius:4px;"></div>
          <div style="position:absolute;top:0;left:0;width:100%;height:100%;
                      display:flex;align-items:center;justify-content:center;
                      color:#5A4A2F;font-weight:bold;">{v}%</div>
        </div>
        """,
        unsafe_allow_html=True
    )

    # ─── se ainda não terminou ───────────────────────────────────
    if v < 100:
        col_back, col_cancel = st.columns(2)

        with col_back:
            if st.button("🔙 Voltar", key="back_progress"):
                st.session_state["pagina"] = "modo"
                st.rerun()

        with col_cancel:
            if st.button("❌ Cancelar Revisão", key="cancel_progress"):
                # exclui pasta de saída e arquivos temporários
                pasta = Path(PASTA_SAIDA) / usuario / nome
                if pasta.exists():
                    shutil.rmtree(pasta, ignore_errors=True)

                for f in [STATUS_PATH, LOG_PROCESSADOS, LOG_FALHADOS]:
                    f.unlink(missing_ok=True)

                remove_from_queue(nome)

                # limpa sessão (mas preserva login)
                for k in list(st.session_state.keys()):
                    if k not in ("user",):
                        del st.session_state[k]

                st.session_state["pagina"] = "upload"
                st.rerun()

            time.sleep(1)
            st.rerun()

    # ────────────────────────── quando chegar a 100 % ─────────────────────
    st.success("✅ Revisão concluída!")

    src_dir = Path(PASTA_SAIDA) / usuario / nome
    doc_final = src_dir / (
        f"{nome}_revisado_texto.docx" if lite else f"{nome}_revisado_completo.docx"
    )
    rel_path = src_dir / f"relatorio_tecnico_{nome}.docx"

    # Espera até 15 s pelo documento final aparecer
    for _ in range(15):
        if doc_final.exists():
            break
        time.sleep(1)
    if not doc_final.exists():
        st.error(f"Arquivo final não encontrado: {doc_final}")
        return

    st.success("✅ Revisão concluída!")

    # upload Drive
    if not st.session_state.get("revision_logged", False):
        user = st.session_state["user"]

        try:
            link_doc = upload_e_link(doc_final)
        except FileNotFoundError as e:
            st.error(str(e))
            return
        link_rel = upload_e_link(rel_path) if rel_path.exists() else None

        ts_now = datetime.now().isoformat()
        log_revision(user["id"], nome, link_doc, ts_now)
        if link_rel:
            log_revision(user["id"], f"Relatório {nome}", link_rel, ts_now)

        backup_db()
        st.session_state["revision_logged"] = True

    st.session_state["pagina"] = "resultados"
    st.rerun()

# PASTA_SAIDA, PASTA_HISTORICO, STATUS_PATH, LOG_PROCESSADOS, LOG_FALHADOS,
# remove_from_queue(), log_revision()  ➜  já existem no seu script

def page_results():
    # 🚫 Se dados básicos faltarem, volta para upload
    user = st.session_state.get("user")
    nome = st.session_state.get("nome")
    usuario = user["username"] if user else st.session_state.get("usuario")
    lite = st.session_state.get("modo_lite", False)

    if not (nome and usuario):
        st.session_state["pagina"] = "upload"
        st.rerun()

    # Remove da fila na primeira renderização
    if not st.session_state.get("removed_from_queue", False):
        remove_from_queue(nome)
        st.session_state["removed_from_queue"] = True

    # --- Caminhos padrão ---------------------------------------------------
    src_dir = Path(PASTA_SAIDA) / usuario / nome
    xlsx = src_dir / "avaliacao_completa.xlsx"
    tokens = src_dir / "mapeamento_tokens.xlsx"

    # --- Espera até 30 s pelo .xlsx ----------------------------------------
    for _ in range(30):
        if xlsx.exists():
            break
        with st.spinner("Processando… aguarde alguns segundos."):
            time.sleep(1)
        st.rerun()

    # Procura alternativa se não encontrou
    if not xlsx.exists():
        possiveis = list(Path(PASTA_SAIDA).glob(f"*/{nome}/avaliacao_completa.xlsx"))
        if possiveis:
            xlsx = possiveis[0]
            src_dir = xlsx.parent
            tokens = src_dir / "mapeamento_tokens.xlsx"

    if not xlsx.exists():
        st.error("Nenhum resultado encontrado em **PASTA_SAIDA**.")
        st.stop()

    # ----------------------------------------------------------------------
    wb = load_workbook(xlsx, data_only=True)
    rs = wb["Resumo"]

    tempo = in_tk = out_tk = 0
    for r in rs.iter_rows(min_row=2, values_only=True):
        if r and len(r) >= 4:
            tempo += r[1] or 0
            in_tk += r[2] or 0
            out_tk += r[3] or 0

    # Tokens adicionais
    if tokens.exists():
        try:
            wb_tok = load_workbook(tokens, data_only=True)
            aba = wb_tok["MapaTokens"]
            for r in aba.iter_rows(min_row=2, values_only=True):
                if r and r[1] and r[2]:
                    in_tk += int(r[1])
                    out_tk += int(r[2])
        except Exception as e:
            st.warning(f"Erro ao somar tokens do mapeador: {e}")

    # Totais por tipo
    tot = {}
    if "Texto" in wb.sheetnames: tot["Textual"] = wb["Texto"].max_row - 1
    if "Bibliográfica" in wb.sheetnames: tot["Bibliográfica"] = wb["Bibliográfica"].max_row - 1
    if "Falhas" in wb.sheetnames: tot["Estrutura"] = wb["Falhas"].max_row - 1

    df = pd.DataFrame.from_dict(tot, orient="index", columns=["Total"]).sort_values("Total")
    cores = {"Textual": "#007f56", "Bibliográfica": "#5A4A2F", "Estrutura": "#00AF74"}

    c1, c2, c3 = st.columns([1, 1.2, 1])

    # 📊 Barras
    with c1:
        st.plotly_chart(
            px.bar(df, orientation="h", color=df.index,
                   color_discrete_map=cores,
                   labels={"index": "Tipo", "Total": "Qtd"}),
            use_container_width=True
        )

    # ─────────── Links do Drive (doc + relatório) ──────────────
    conn = sqlite3.connect(DB_PATH)
    rows_link = conn.execute(
        """
        SELECT file_name, processed_path
        FROM revisions
        WHERE user_id = ?
          AND (file_name = ? OR file_name = ?)
        ORDER BY timestamp DESC
        """,
        (user["id"], nome, f"Relatório {nome}")
    ).fetchall()
    conn.close()

    link_doc = link_rel = None
    for fn, pth in rows_link:
        if pth.startswith(("http://", "https://")):
            if fn.startswith("Relatório"):
                link_rel = pth
            else:
                link_doc = pth

    # helper
    def prefer_local_then_link(path_local: Path, link_drive: str | None):
        if path_local.exists():
            return ("local", path_local)
        if link_drive:
            return ("link", link_drive)
        return (None, None)

    # 📈 Métricas + Downloads
    with c2:
        st.metric("⏱ Tempo (s)", f"{tempo:.1f}")
        st.metric("📝 Palavras de Entrada", f"{int(in_tk)*0.75:.0f}")
        st.metric("✍️ Palavras Alteradas", f"{int(out_tk)*0.75:.0f}")

        docs = [(f"{nome}_revisado_texto.docx", "📄 Documento Revisado (Lite)", link_doc)] if lite else \
               [(f"{nome}_revisado_completo.docx", "📄 Documento Revisado", link_doc)]
        docs.append((f"relatorio_tecnico_{nome}.docx", "📑 Relatório Técnico", link_rel))

        for fn, lbl, link in docs:
            origem, recurso = prefer_local_then_link(src_dir / fn, link)
            if origem == "local":
                st.download_button(lbl, recurso.read_bytes(),
                                   file_name=recurso.name,
                                   key=f"dl_{fn}")
            elif origem == "link":
                st.link_button(lbl, url=recurso, use_container_width=True)

    # 🥧 Pizza
    with c3:
        st.plotly_chart(
            px.pie(df, values="Total", names=df.index,
                   color_discrete_map=cores,
                   title="Distribuição %"),
            use_container_width=True
        )
    st.markdown("---")
    if st.button("🔙 Voltar", key="back_results"):
        st.session_state["pagina"] = "upload"
        st.rerun()


# --- Footer ---
def footer():
    st.markdown('<hr style="margin-top: 2rem; margin-bottom: 1rem; border: none; border-top: 1px solid #ccc;"/>', unsafe_allow_html=True)
    st.markdown('<div class="footer" style="color: #007f56; font-weight: bold;">Powered by Dossel Ambiental</div>', unsafe_allow_html=True)


# --- Main ---------------------------------------------------------------
st.set_page_config(page_title='Revisor Dossel', layout='centered')

def main():
    # 1️⃣ Restaura o banco apenas se as variáveis da service-account existirem
    try:
        if "SA_KEY_B64" in os.environ and "FOLDER_ID" in os.environ:
            restore_db()       # traz users.db do Drive
    except Exception as e:
        print("[restore_db] erro ignorado ➜", e)

    init_db()
    apply_css()

    if "pagina" not in st.session_state:
        st.session_state["pagina"] = "upload" if "user" in st.session_state else "login"

    # 🔐 Se não logado, força página de login
    if "user" not in st.session_state:
        if st.session_state["pagina"] != "login":
            st.session_state["pagina"] = "login"
            st.rerun()
        header()
        page_login()
        footer()
        return

    # === SIDEBAR ========================================================
    with st.sidebar:
        pagina_atual = st.session_state.get("pagina", "upload")
        index_padrao = 1 if pagina_atual == "historico" else 0

        secao = option_menu(
            menu_title=None,
            options=["Nova Revisão", "Histórico"],
            icons=["file-earmark-text", "clock-history"],
            default_index=index_padrao,
            styles={
                "container": {"padding": "0!important", "background-color": "transparent"},
                "icon": {"color": "#00AF74", "font-size": "18px"},
                "nav-link": {"margin": "2px 0", "--hover-color": "#f7f7f7"},
                "nav-link-selected": {"background-color": "#00AF74"},
            }
        )

        # guarda página anterior para voltar da aba Histórico
        if st.session_state["pagina"] not in ["historico", "login"]:
            st.session_state["pagina_revisao"] = st.session_state["pagina"]

        if secao == "Histórico" and st.session_state["pagina"] != "historico":
            st.session_state["pagina"] = "historico"
            st.rerun()
        elif secao == "Nova Revisão":
            voltar = st.session_state.get("pagina_revisao", "upload")
            if st.session_state["pagina"] != voltar:
                st.session_state["pagina"] = voltar
                st.rerun()

        # 🔘 Logout
        if st.button("❌ Logout (sair)", use_container_width=True):
            nome = st.session_state.get("nome")
            if nome:
                pasta = Path(PASTA_SAIDA) / st.session_state['usuario'] / nome
                if pasta.exists():
                    shutil.rmtree(pasta)
            for f in ["status.txt", "documentos_processados.txt", "documentos_falhados.txt"]:
                Path(f).unlink(missing_ok=True)
            remove_from_queue(nome)

            # 2️⃣ Faz backup imediato do banco antes de limpar a sessão
            try:
                if "SA_KEY_B64" in os.environ and "FOLDER_ID" in os.environ:
                    backup_db()
            except Exception as e:
                print("[backup_db] erro ignorado ➜", e)

            st.session_state.clear()
            st.rerun()

    # === CONTEÚDO PRINCIPAL ============================================
    header()
    pagina = st.session_state["pagina"]

    if pagina == "login":
        page_login()
    elif pagina == "upload":
        page_upload()
    elif pagina == "modo":
        page_mode()
    elif pagina == "acompanhamento":
        page_progress()
    elif pagina == "resultados":
        page_results()
    elif pagina == "historico":
        page_history()
    else:
        st.error("Página inválida")

    footer()

if __name__ == "__main__":
    main()
